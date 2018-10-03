# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from order import Order
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import time
from login import Login


def readFile(fileName):
    wb = load_workbook(filename=fileName, read_only=True)
    ws = wb['Sheet1']
    listLogin = []

    for row in tuple(ws.rows)[1:]:
        if(row[0].value and row[1].value is not None):
            register = Login(row[0].value, row[1].value)
            listLogin.append(register)
    return listLogin


def writeFile(data):
    # print('data=',data)
    wb = Workbook()
    ws = wb.active
    ws.append(['Họ và tên', 'Số điện thoại',  'Trạng thái',
               'Giá sản phẩm', 'Địa chỉ', 'Mã đơn hàng', 'Email', 'Link check'])
    for row in data:
        ws.append(row)
    wb.save('out.xlsx')


def run():
    listLogin = readFile('login.xlsx')
    listInfo = []
    for element in listLogin:
        info = []
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--incognito")
        chrome_options.add_argument("--start-maximized")
        browser = webdriver.Chrome(
            './chromedriver', chrome_options=chrome_options)
        link = 'https://www.lazada.vn'
        browser.get(link)
        browser.find_element_by_xpath("//div[@id='anonLogin']//a").click()
        browser.find_element_by_xpath(
            "//div[@class='mod-input mod-login-input-loginName mod-input-loginName']//input").send_keys(element.email)
        browser.find_element_by_xpath(
            "//div[@class='mod-input mod-input-password mod-login-input-password mod-input-password']//input").send_keys(element.password)
        browser.find_element_by_xpath(
            "//div[@class='mod-login-btn']//button").click()
        time.sleep(1)
        link = "https://my.lazada.vn/customer/order/index/?spm=a2o4n"
        browser.get(link)
        # print('text',browser.find_element_by_xpath("//div[@class='order-info']").text)
        orderText = browser.find_element_by_xpath(
            "//div[@class='order-info']//a[@class='link']").text
        orderTextSplit = orderText.split("#")[1]
        print(orderTextSplit)
        browser.find_element_by_xpath(
            "//div[@class='order-info']//a[@class='link']").click()
        totalPrice = browser.find_element_by_xpath(
            "//span[@class='text bold total-price pull-right']").text
        delivery = browser.find_element_by_class_name("delivery-wrapper")
        temp = delivery.text.splitlines()
        name = temp[1]
        address = temp[2]
        phone = temp[3]
        # print('....', temp)

        try:
            status = browser.find_element_by_class_name(
                'tracking-item-content').text
        except NoSuchElementException:
            status = 'Đã hủy'
        link = 'https://my.lazada.vn/customer/order/view/?tradeOrderId=' + \
            orderTextSplit + '&buyerEmail=' + element.email
        info.append(name)
        info.extend([phone, status, totalPrice, address,
                     orderTextSplit, element.email, link])
        # info = Info(name, address, phone, link, element.orderID, element.email, status, totalPrice)
        listInfo.append(info)

        browser.quit()

        writeFile(listInfo)


run()
