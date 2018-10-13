# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from order import Order
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import time
from threading import Thread
import threading
from login import Login


def readFile(fileName):
    wb = load_workbook(filename=fileName, read_only=True)
    ws = wb['Sheet1']
    listLogin = []

    for row in tuple(ws.rows)[1:]:
        if(row[0].value and row[1].value is not None):
            register = Login(row[0].value, row[1].value)
            listLogin.append(register)
    return listLogin


def writeFile(data):
    # print('data=',data)
    wb = Workbook()
    ws = wb.active
    ws.append(['Họ và tên', 'Số điện thoại',  'Trạng thái',
               'Giá sản phẩm', 'Địa chỉ', 'Mã đơn hàng', 'Email', 'Link check'])
    for row in data:
        ws.append(row)
    wb.save('out.xlsx')


def browserMultiThread(email, password, slashlink):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("--start-maximized")
    browser = webdriver.Chrome(
        './chromedriver', chrome_options=chrome_options)
    link = 'https://www.lazada.vn'
    browser.get(link)
    browser.find_element_by_xpath("//div[@id='anonLogin']//a").click()
    browser.find_element_by_xpath(
        "//div[@class='mod-input mod-login-input-loginName mod-input-loginName']//input").send_keys(email)
    browser.find_element_by_xpath(
        "//div[@class='mod-input mod-input-password mod-login-input-password mod-input-password']//input").send_keys(password)
    print('sdt: ', email)

    browser.find_element_by_xpath(
        "//div[@class='mod-login-btn']//button").click()
    time.sleep(1)
    browser.get(slashlink)
    browser.find_element_by_css_selector(
        "#scroller_rv > div > div:nth-child(3) > div:nth-child(1) > div:nth-child(4)").click()
    browser.quit()


def run():
    listLogin = readFile('login.xlsx')
    link = "https://pages.lazada.vn/wow/i/vn/slash-it/sla?wh_sellerId=100101649&wh_promotionId=202147000301649&wh_interactActivityId=1208090&wh_weex=true&wh_language=vi"
    try:
        for element in listLogin:
            t = time.time()
            t1 = threading.Thread(
                target=browserMultiThread, args=(element.email, element.password, link))
            t2 = threading.Thread(
                target=browserMultiThread, args=(element.email, element.password, link))
            t3 = threading.Thread(
                target=browserMultiThread, args=(element.email, element.password, link))
            t1.start()
            # t2.start()
            # t3.start()
            t1.join()
            # t2.join()
            # t3.join()
            print("done in ", time.time() - t)
    except:
        print("error")


run()
