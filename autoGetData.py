from openpyxl import load_workbook
from openpyxl import Workbook
from order import Order
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import time
def readFile(fileName):
    wb = load_workbook(filename=fileName, read_only=True)
    ws = wb['Sheet1']
    listOrder = []

    # orders = Order()
    # rowValues = ws.rows.value
    # orders['orderID'] = rowValues[0]
    # orders['email'] = rowValues[1]
    # print(ws.rows[1:])
    for row in tuple(ws.rows)[1:]:
        if(row[0].value and row[1].value is not None):
            orders = Order(row[0].value, row[1].value)
            listOrder.append(orders)

    # for order in listOrder:
    #     print('...', order.orderID, order.email)
    return listOrder


def writeFile(data):
    # print('data=',data)
    wb = Workbook()
    ws = wb.active
    ws.append(['Họ và tên', 'Số điện thoại',  'Trạng thái', 'Giá sản phẩm','Địa chỉ','Mã đơn hàng', 'Email', 'Link check'])
    for row in data:
        ws.append(row)
    wb.save('out.xlsx')
def writeError(data):
    # print('data=',data)
    wb = Workbook()
    ws = wb.active
    ws.append(['Họ và tên', 'Số điện thoại',  'Trạng thái', 'Giá sản phẩm','Địa chỉ','Mã đơn hàng', 'Email', 'Link check'])
    for row in data:
        ws.append(row)
    wb.save('error.xlsx')
def run():
    listOrder = readFile('data.xlsx')
    # browser = webdriver.Chrome(executable_path=)
    # browser = webdriver.Chrome('./chromedriver')
    listInfo = []
    for element in listOrder:
        browser = webdriver.Chrome('./chromedriver')
        info = []
        print('...', element.orderID, element.email)
        link = 'https://my.lazada.vn/customer/order/view/?tradeOrderId=' + str(element.orderID) + '&buyerEmail=' + element.email
        try:
            browser.get(link)
            time.sleep(1)
            totalPrice = browser.find_element_by_xpath("//span[@class='text bold total-price pull-right']").text
            delivery = browser.find_element_by_class_name("delivery-wrapper")
            temp = delivery.text.splitlines()
            name = temp[1]
            address = temp[2]
            phone = temp[3]
            # print('....', temp)
            
            try:
                status = browser.find_element_by_class_name('tracking-item-content').text
            except NoSuchElementException:
                status = 'Đã hủy'

            info.append(name)
            info.extend([phone, status, totalPrice, address, element.orderID,element.email, link])
            # info = Info(name, address, phone, link, element.orderID, element.email, status, totalPrice)
            listInfo.append(info)
            writeFile(listInfo)
        except NoSuchElementException: 
            writeError(listInfo)
    browser.quit()
        # for element in listInfo:
        # print('...', element.name, element.address, element.phone, element.orderID, element.email, element.status, element.totalPrice)
    
run()
# writeFile([])
