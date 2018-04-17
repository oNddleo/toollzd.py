from openpyxl import load_workbook
from openpyxl import Workbook
from order import Order
from selenium import webdriver


def readFile(nameFile):
    wb = load_workbook(filename=nameFile, read_only=True)
    ws = wb['Sheet1']
    listOrder = []

    # orders = Order()
    # rowValues = ws.rows.value
    # orders['orderID'] = rowValues[0]
    # orders['email'] = rowValues[1]
    # print(ws.rows[1:])
    for row in tuple(ws.rows)[1:]:
        if(row[0].value and row[1].value is not None):
            orders = Order(row[0].value, row[1].value)
            listOrder.append(orders)

    # for order in listOrder:
    #     print('...', order.orderID, order.email)
    return listOrder


def writeFile(data):
    # print('data=',data)
    wb = Workbook()
    ws = wb.active
    ws.append(['Email','Mã đơn hàng', 'Link check', 'Trạng thái', 'Họ và tên', 'Địa chỉ', 'Số điện thoại', 'Giá sản phẩm'])
    for row in data:
        ws.append(row)
    wb.save('out.xlsx')

def run():
    listOrder = readFile('data.xlsx')
    # browser = webdriver.Chrome(executable_path=)
    browser = webdriver.Chrome('./chromedriver')
    listInfo = []
    for element in listOrder:
        info = []
        print('...', element.orderID, element.email)
        link = 'https://my.lazada.vn/customer/order/view/?tradeOrderId=' + str(element.orderID) + '&buyerEmail=' + element.email
        browser.get(link)
        totalPrice = browser.find_element_by_class_name('total-price').text
        delivery = browser.find_element_by_class_name('delivery-wrapper')
        temp = delivery.text.splitlines()
        name = temp[1]
        address = temp[2]
        phone = temp[3]
        # print('....', temp)
        
        status = browser.find_element_by_class_name('tracking-item-content').text
        info.append(element.email)
        info.extend([element.orderID, link, status, name, address, phone, totalPrice])
        # info = Info(name, address, phone, link, element.orderID, element.email, status, totalPrice)
        listInfo.append(info)
        
    browser.quit()
    # for element in listInfo:
        # print('...', element.name, element.address, element.phone, element.orderID, element.email, element.status, element.totalPrice)
    writeFile(listInfo)
run()
# writeFile([])
